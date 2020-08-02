import shutil
import datetime
import openpyxl
import os
import os.path
import win32com.client
from excelDbManager.errors import *

xlOpenXMLWorkbookMacroEnabled = 52

# Sheets
S_BASE = "Base"
S_MANAGE = "Manage"
S_MULTIPLE = "Ajouts Multiple"

# Cells
CELLS_PUBLIC = {'id_base': 'B2', 'id_personne': 'C2', 'domaine_etude': 'D2', 'pays_etude': 'E2', 'etablissement': 'F2',
                'parcours_com': 'G2', 'domaine_pro': 'H2', 'metier': 'I2', 'employeur': 'J2', 'pays_pro': 'K2',
                'com': 'L2'}
CELLS_PRIVATE = {'id_personne': 1, 'nom': 2, 'prenom': 3, 'email': 4, 'tel': 5, 'linkedin': 6, 'situation': 7,
                 'promo': 8, 'date': 9}


def create_backup(source, destination):
    """
    :param source: path to the file to copy
    :param destination: path to the back up folder
    :return: create a backup of source in destination
    """
    now = datetime.datetime.now()
    date = now.strftime("%Y_%m_%d__%Hh%M")
    name, extension = get_name_from_path(source)
    destination = destination + "/" + name + "_" + date + "." + extension
    try:
        shutil.copy(source, destination)
        backup = True
    except:
        backup = False
    return backup


def get_name_from_path(path):
    """
    :param path:
    :return: extract name and extension of
    """
    sep = path.split(sep='/')
    name_with_extension = sep[len(sep) - 1]
    name, extension = name_with_extension.split(".")
    return name, extension


def exist_in_db_id(id, source):
    """
    :param id: int
    :param source: path to database
    :return: True if id exist in source else False
    """
    wb = openpyxl.load_workbook(filename=source)
    sheet = wb[S_BASE]
    row = 3
    while (sheet['A' + str(row)].value is not None) and (sheet['A' + str(row)].value != id):
        row += 1
    if sheet['A' + str(row)].value is None:
        exist = False
    else:
        exist = True
    return exist


def get_first_empty_row(sheet):
    primary_key = -1
    row = 3
    while sheet.Cells(row, 1).value is not None:
        primary_key = max(primary_key, sheet.Cells(row, 1).value)
        row += 1
    return row, primary_key + 1


def add_line_public(data, path):
    wb = openpyxl.load_workbook(filename=path)
    sheet = wb[S_MANAGE]
    # Place data
    for key, value in data:
        sheet[CELLS_PUBLIC[key]].value = value
    # save to take change into account
    wb.save(path)
    # activate maccro
    run_maccro(path, 'Ajouter_ligne_python')
    return -1


def add_line_private(data, path):
    xl = win32com.client.Dispatch("Excel.Application")  # activate excel
    wb = xl.Workbooks.Open(os.path.abspath(path))  # open workbook
    xlSheet = wb.Sheets(S_BASE)  # select sheet
    # find where to write
    row, primary_key = get_first_empty_row(xlSheet)

    # add data
    xlSheet.Cells(row, CELLS_PRIVATE['id_personne']).value = primary_key
    xlSheet.Cells(row, CELLS_PRIVATE['date']).value = '=NOW()'
    for element in data:
        if (element != 'id_personne') and (element != 'date'):
            xlSheet.Cells(row, CELLS_PRIVATE[element]).value = data[element]

    # save and quit
    xl.DisplayAlerts = False  # avoid the alert about replacing the file
    wb.SaveAs(os.path.abspath(path), FileFormat=xlOpenXMLWorkbookMacroEnabled)  # save file with macros
    xl.Quit()

    return primary_key


def modif_line_public(data, path):
    wb = openpyxl.load_workbook(filename=path)
    sheet = wb[S_MANAGE]
    # Place data
    for key, value in data:
        sheet[CELLS_PUBLIC[key]].value = value
    # save to take change into account
    wb.save(path)
    # activate maccro
    run_maccro(path, 'Modifier_ligne_python')
    return -1


def modif_line_private(data, path):
    wb = openpyxl.load_workbook(filename=path)
    sheet = wb[S_MANAGE]
    # Place data
    for key, value in data:
        sheet[CELLS_PRIVATE[key]].value = value
    # save to take change into account
    wb.save(path)
    # activate maccro
    run_maccro(path, 'Modifier_ligne_python')
    return -1
